from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
from openpyxl import Workbook

from src.database.conexiones import get_db
from src.reports.adolescentes import (
    total_adolescentes,
    adolescentes_por_categoria,
    adolescentes_por_institucion,
    adolescentes_por_actividad,
    top10_instituciones,
    top10_actividades,
    adolescentes_por_tramo_edad,
    adolescentes_por_genero
)

router = APIRouter(prefix="/reportes-excel", tags=["Reportes Excel"])


# ---------------------------------------------
# FUNCIÓN AUXILIAR PARA GENERAR ARCHIVO EXCEL
# ---------------------------------------------
def crear_excel(nombre_hoja: str, encabezados: list, filas: list):
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    ws.append(encabezados)

    for fila in filas:
        ws.append(list(fila))

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ---------------------------------------------
# REPORTES INDIVIDUALES
# ---------------------------------------------

@router.get("/categoria")
def excel_categoria(db: Session = Depends(get_db)):
    datos = adolescentes_por_categoria(db)
    excel = crear_excel(
        "Categorías",
        ["Categoría", "Cantidad"],
        datos
    )
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=categorias.xlsx"}
    )


@router.get("/instituciones")
def excel_instituciones(db: Session = Depends(get_db)):
    datos = adolescentes_por_institucion(db)
    excel = crear_excel(
        "Instituciones",
        ["Institución", "Cantidad"],
        datos
    )
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=instituciones.xlsx"}
    )


@router.get("/actividades")
def excel_actividades(db: Session = Depends(get_db)):
    datos = adolescentes_por_actividad(db)
    excel = crear_excel(
        "Actividades",
        ["Actividad", "Cantidad"],
        datos
    )
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=actividades.xlsx"}
    )


@router.get("/top10-instituciones")
def excel_top10_instituciones(db: Session = Depends(get_db)):
    datos = top10_instituciones(db)
    excel = crear_excel(
        "Top 10 Instituciones",
        ["Institución", "Cantidad"],
        datos
    )
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=top10_instituciones.xlsx"}
    )


@router.get("/top10-actividades")
def excel_top10_actividades(db: Session = Depends(get_db)):
    datos = top10_actividades(db)
    excel = crear_excel(
        "Top 10 Actividades",
        ["Actividad", "Cantidad"],
        datos
    )
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=top10_actividades.xlsx"}
    )


@router.get("/tramo-edad")
def excel_tramo_edad(db: Session = Depends(get_db)):
    datos = adolescentes_por_tramo_edad(db)
    excel = crear_excel(
        "Tramo Edad",
        ["Tramo Edad", "Cantidad"],
        datos
    )
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tramo_edad.xlsx"}
    )


@router.get("/genero")
def excel_genero(db: Session = Depends(get_db)):
    datos = adolescentes_por_genero(db)
    excel = crear_excel(
        "Género",
        ["Género", "Cantidad"],
        datos
    )
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=genero.xlsx"}
    )


# ----------------------------------------------------
# OPCIONAL: TODO JUNTO EN UN SOLO ARCHIVO MULTI-HOJA
# ----------------------------------------------------
@router.get("/completo")
def excel_completo(db: Session = Depends(get_db)):
    wb = Workbook()

    # Hoja 1 — Categoría
    ws = wb.active
    ws.title = "Categoría"
    ws.append(["Categoría", "Cantidad"])
    for fila in adolescentes_por_categoria(db):
        ws.append(list(fila))

    # Hoja 2 — Institución
    ws2 = wb.create_sheet("Institución")
    ws2.append(["Institución", "Cantidad"])
    for fila in adolescentes_por_institucion(db):
        ws2.append(list(fila))

    # Hoja 3 — Actividades
    ws3 = wb.create_sheet("Actividad")
    ws3.append(["Actividad", "Cantidad"])
    for fila in adolescentes_por_actividad(db):
        ws3.append(list(fila))

    # Hoja 4 — Tramo Edad
    ws4 = wb.create_sheet("Edad")
    ws4.append(["Tramo Edad", "Cantidad"])
    for fila in adolescentes_por_tramo_edad(db):
        ws4.append(list(fila))

    # Hoja 5 — Género
    ws5 = wb.create_sheet("Género")
    ws5.append(["Género", "Cantidad"])
    for fila in adolescentes_por_genero(db):
        ws5.append(list(fila))

    # Guardar
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_completo.xlsx"}
    )
